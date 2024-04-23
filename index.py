import time
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook
from selenium import webdriver
from tkinter import messagebox
from selenium.webdriver.chrome.service import Service as ChromeService, Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class ExcelURLOpener:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel URL Opener")
        self.root.geometry("400x200")
        self.style = ttk.Style()
        self.style.configure('TButton', foreground='#000000', background='#007bff', font=('Helvetica', 10))
        self.style.map('TButton', foreground=[('active', '#ffffff'), ('pressed', '#ffffff')], background=[('active', '#0056b3'), ('pressed', '#0056b3')])
        self.style.configure('TLabel', foreground='#333333', background='#ffffff', font=('Helvetica', 10))
        self.style.configure('TFrame', background='#ffffff')
        self.upload_button = ttk.Button(self.root, text="Upload Excel File", command=self.open_excel_file)
        self.upload_button.pack(pady=20)
        self.run_button = ttk.Button(self.root, text="Run WebDriver", command=self.run_webdriver)
        self.run_button.pack(pady=10)

    def open_excel_file(self):
        filename = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if filename:
            self.read_excel_file(filename)

    def read_excel_file(self, filename):
        try:
            self.urls = []
            wb = load_workbook(filename)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                url = row[0]  # Assuming URL is in the first column
                if url is not None and url.startswith("http"):
                    self.urls.append(url)
            if not self.urls:
                messagebox.showwarning("Warning", "No valid URLs found in the Excel file.")
            else:
                messagebox.showinfo("Success", "Excel file successfully loaded.")
        except Exception as e:
            messagebox.showerror("Error", "Error occurred while reading Excel file.")
            print("Error occurred while reading Excel file:", e)

    def run_webdriver(self):
        if hasattr(self, 'urls') and self.urls:
            try:
                driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
                driver.implicitly_wait(10)
                for url in self.urls:

                    driver.get(url)
                    driver.implicitly_wait(10)
                    driver.find_element(By.XPATH,"//button[@class='cb b cc cd ce cg ch ci cj ck cl cm cn co cp cq cr cs ct cu cv cw cx cy cz']").click()
                    time.sleep(10)
                    driver.find_element(By.XPATH, "//div[text()='Sign in with email']").click()
                    email_id = driver.find_element(By.CSS_SELECTOR,"input[class='re km rf rg rh ri rj rk rl rm rn ro rd rp rq rr cz ce rs']").click()
                    email_id.send_key("abhiraz2019@gmail.com")
                    continue_btn = driver.find_element(By.XPATH,
                                        "//button[@class='cb b cc cd ce cg ch ci cj ck cl cm cn co cp cq cr cs ct cu cv cw cx cy cz']")
                    driver.get_cookies()

            except Exception as e:
                messagebox.showerror("Error", "Error occurred while opening URLs.")
                print("Error occurred while opening URLs:", e)
            finally:
                driver.quit()
        else:
            messagebox.showwarning("Warning", "Please upload an Excel file first.")

# Create GUI
root = tk.Tk()
app = ExcelURLOpener(root)
root.mainloop()
